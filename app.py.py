import io
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

st.set_page_config(page_title="盤點資料自動彙整", layout="wide")

st.title("📦 庫存盤點自動彙整系統")
st.write("請將各單位填寫完畢的 Excel 盤點檔上傳，系統會自動幫您合併並計算總表。")
st.info("💡 追蹤秘訣：系統會自動將「檔案名稱-分頁名稱」當作來源，建議同仁將檔名取為【調劑台名稱_姓名】。")

st.warning("⚠️ 注意：上傳的 Excel 檔案裡面，第一列的表頭必須包含這三個名稱：【料位號】、【藥品名】、【數量】（前後多餘空白會自動忽略）")

REQUIRED_COLS = ["料位號", "藥品名", "數量"]


# ----------------------------------------------------------------------
# 資料清洗函式
# ----------------------------------------------------------------------
def clean_code_column(series: pd.Series) -> pd.Series:
    """
    清洗料位號 / 代碼類欄位。
    避免數字欄位被讀成 float 後變成 "101.0" 而非 "101"，
    導致同一個料位在不同分頁被誤判為不同群組。
    """
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.notna().mean() > 0.9:
        cleaned = numeric.astype("Int64").astype(str)
        fallback = series.astype(str)
        cleaned = cleaned.where(numeric.notna(), fallback)
    else:
        cleaned = series.astype(str)

    cleaned = (
        cleaned.str.replace(r"[\r\n]+", "", regex=True)
        .str.replace("_x000D_", "", regex=False)
        .str.strip()
    )
    return cleaned


def clean_text_column(series: pd.Series) -> pd.Series:
    """清洗藥品名等文字欄位：去除換行、Excel 殘留字元、多餘空白（含全形空格）。"""
    cleaned = (
        series.astype(str)
        .str.replace(r"[\r\n]+", "", regex=True)
        .str.replace("_x000D_", "", regex=False)
        .str.replace("\u3000", " ", regex=False)
        .str.strip()
    )
    cleaned = cleaned.str.replace(r"\s+", " ", regex=True)
    return cleaned


# ----------------------------------------------------------------------
# 記憶體優化讀取：openpyxl read_only 模式，逐列串流讀取
# 只抓需要的 3 個欄位，不把整張表（含其他欄位、格式、樣式）都載入記憶體
# ----------------------------------------------------------------------
def read_xlsx_sheets(file_bytes: bytes):
    """
    回傳 list of (sheet_name, dataframe_or_None, header_list)
    dataframe 為 None 代表這個分頁缺少必備欄位，header_list 供錯誤訊息顯示用。
    """
    results = []
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)

            try:
                header_row = next(rows_iter)
            except StopIteration:
                results.append((sheet_name, None, ["(空白分頁)"]))
                continue

            header = [str(h).strip() if h is not None else "" for h in header_row]
            col_index = {name: idx for idx, name in enumerate(header) if name in REQUIRED_COLS}

            if not all(c in col_index for c in REQUIRED_COLS):
                results.append((sheet_name, None, header))
                continue

            records = []
            for row in rows_iter:
                try:
                    code = row[col_index["料位號"]]
                    name = row[col_index["藥品名"]]
                    qty = row[col_index["數量"]]
                except IndexError:
                    continue
                if code is None and name is None and qty is None:
                    continue  # 整列空白直接跳過
                records.append((code, name, qty))

            if records:
                df = pd.DataFrame.from_records(records, columns=REQUIRED_COLS)
            else:
                df = pd.DataFrame(columns=REQUIRED_COLS)

            results.append((sheet_name, df, header))
    finally:
        wb.close()  # 明確釋放，避免大檔案處理完後仍佔用記憶體

    return results


def read_xls_sheets(file_bytes: bytes):
    """
    舊版 .xls 格式 openpyxl 無法開啟，退回用 pandas + xlrd。
    .xls 本身有 65536 列上限，通常檔案不會太大，記憶體風險相對低。
    """
    try:
        xls_dict = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, engine="xlrd")
    except ImportError:
        raise RuntimeError("讀取 .xls 需要安裝 xlrd 套件（pip install xlrd），建議請同仁改存成 .xlsx")

    results = []
    for sheet_name, df in xls_dict.items():
        df.columns = [str(c).strip() for c in df.columns]
        if not all(col in df.columns for col in REQUIRED_COLS):
            results.append((sheet_name, None, list(df.columns)))
            continue
        results.append((sheet_name, df[REQUIRED_COLS].copy(), list(df.columns)))
    return results


# ----------------------------------------------------------------------
# 主流程
# ----------------------------------------------------------------------
uploaded_files = st.file_uploader(
    "📂 批次上傳所有 Excel 盤點檔", type=["xlsx", "xls"], accept_multiple_files=True
)

if uploaded_files:
    all_data = []
    error_files = []      # (檔名, 原因)
    skipped_sheets = []   # (檔名, 分頁, 實際欄位)

    progress = st.progress(0.0, text="開始處理檔案…")

    for i, file in enumerate(uploaded_files):
        progress.progress(i / len(uploaded_files), text=f"處理中：{file.name}")

        file_bytes = file.read()  # 讀一次成 bytes，之後不再依賴上傳物件本身的游標狀態

        try:
            if file.name.lower().endswith(".xls"):
                sheets_data = read_xls_sheets(file_bytes)
            else:
                sheets_data = read_xlsx_sheets(file_bytes)
        except Exception as e:
            error_files.append((file.name, f"讀取失敗：{e}"))
            del file_bytes
            continue

        del file_bytes  # 這個檔案的原始 bytes 已經沒用了，提早釋放

        file_has_valid_sheet = False
        for sheet_name, df, header in sheets_data:
            if df is None:
                skipped_sheets.append((file.name, sheet_name, ", ".join(header)))
                continue
            if df.empty:
                skipped_sheets.append((file.name, sheet_name, "此分頁沒有資料列"))
                continue

            df["藥品名"] = clean_text_column(df["藥品名"])
            df["料位號"] = clean_code_column(df["料位號"])
            df["來源檔名 (分頁)"] = f"{file.name} - {sheet_name}"

            all_data.append(df)
            file_has_valid_sheet = True

        if not file_has_valid_sheet:
            error_files.append((file.name, "所有分頁皆缺少必備欄位（料位號/藥品名/數量）或無資料"))

        del sheets_data

    progress.progress(1.0, text="處理完成")
    progress.empty()

    # --- 錯誤與略過的分頁提示 ---
    if error_files:
        st.error("❌ 下列檔案讀取失敗或完全沒有必備欄位：")
        for fname, reason in error_files:
            st.write(f"　• **{fname}**：{reason}")

    if skipped_sheets:
        with st.expander(f"⚠️ 有 {len(skipped_sheets)} 個分頁因欄位不符或無資料被略過（點此展開查看）"):
            for fname, sheet, cols in skipped_sheets:
                st.write(f"　• **{fname}** - 分頁「{sheet}」，實際欄位：{cols}")

    if all_data:
        master_df = pd.concat(all_data, ignore_index=True)
        del all_data  # 合併完就不需要個別小 DataFrame 了，釋放記憶體

        # --- 數量欄位轉換，並找出無法辨識的資料 ---
        qty_numeric = pd.to_numeric(master_df["數量"], errors="coerce")
        bad_qty_mask = qty_numeric.isna()

        if bad_qty_mask.any():
            st.warning(
                f"⚠️ 有 {bad_qty_mask.sum()} 筆「數量」欄位無法辨識為數字，已暫時視為 0 計算，"
                f"請至下方確認並回原始檔案修正："
            )
            st.dataframe(
                master_df.loc[bad_qty_mask, ["料位號", "藥品名", "數量", "來源檔名 (分頁)"]],
                use_container_width=True,
                hide_index=True,
            )

        master_df["數量"] = qty_numeric.fillna(0)

        # --- 疑似重複品名提醒（同料位號下，藥品名去除所有空白後相同，但原字串不同）---
        dedupe_key = master_df["藥品名"].str.replace(r"\s+", "", regex=True)
        key_series = master_df["料位號"].astype(str) + "|" + dedupe_key
        dup_groups = master_df.groupby(key_series)["藥品名"].nunique()
        suspicious_keys = dup_groups[dup_groups > 1].index

        if len(suspicious_keys) > 0:
            suspicious_rows = master_df[key_series.isin(suspicious_keys)].sort_values(["料位號", "藥品名"])
            with st.expander(f"🔎 發現 {len(suspicious_keys)} 組疑似重複品名（同料位、名稱疑似只差空白/格式），點此查看"):
                st.dataframe(
                    suspicious_rows[["料位號", "藥品名", "數量", "來源檔名 (分頁)"]],
                    use_container_width=True,
                    hide_index=True,
                )

        # --- 製作總表 ---
        source_df = master_df.sort_values(by=["料位號"])

        summary_df = master_df.groupby(["料位號", "藥品名"], as_index=False)["數量"].sum()
        summary_df = summary_df.sort_values(by="料位號")

        st.success(f"✅ 成功合併了 {len(uploaded_files)} 個檔案中的所有分頁！結果如下：")

        tab1, tab2 = st.tabs(["📊 1. 盤點總表 (已加總)", "🔍 2. 追蹤來源明細"])

        with tab1:
            st.subheader("盤點總表")
            st.dataframe(summary_df, use_container_width=True, hide_index=True)

            csv_summary = summary_df.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "📥 下載盤點總表 (CSV)",
                data=csv_summary,
                file_name="1_盤點總表.csv",
                mime="text/csv",
                type="primary",
            )

        with tab2:
            st.subheader("追蹤來源總表 (所有原始明細)")
            st.dataframe(source_df, use_container_width=True, hide_index=True)

            csv_source = source_df.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "📥 下載追蹤來源總表 (CSV)",
                data=csv_source,
                file_name="2_追蹤來源總表.csv",
                mime="text/csv",
            )
